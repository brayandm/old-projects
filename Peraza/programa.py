from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
from openpyxl.styles import Border, Side

def trans(cad):
	temp = ""
	for i in range(len(cad)-1):
		temp += cad[i];
	return temp;

wb = Workbook()

ws1 = wb.active 
ws2 = wb.create_sheet() 
ws3 = wb.create_sheet()
ws4 = wb.create_sheet()

ws1.title = "Sala 1";
ws2.title = "Sala 2";
ws3.title = "Sala 3";
ws4.title = "Sala 4";

def fill(sheet, file):
	
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 6
	sheet.column_dimensions['C'].width = 90

	data = open(file, 'r').readlines()

	for i in range(len(data)):
		data[i] = trans(data[i])

	nombres = []
	edades = []
	diags = []

	for i in range(len(data)):
		if i % 4 == 0:
			nombres.append(data[i]);
		if i % 4 == 1:
			edades.append(data[i]);
		if i % 4 == 2:
			diags.append(data[i]);

	sheet['A1'] = "Nombre y Apellido";
	sheet['B1'] = "Edad";
	sheet['C1'] = "Diagnostico";

	tam = max(len(nombres), len(edades), len(diags))

	for i in range(tam):
		sheet['A' + str(i+2)] = str(nombres[i]);

	for i in range(tam):
		sheet['B' + str(i+2)] = str(edades[i]);

	for i in range(tam):
		sheet['C' + str(i+2)] = str(diags[i]);


fill(ws1,'sala1.txt')
fill(ws3,'sala3.txt')

wb.save("Hogar Peraza.xlsx")
 